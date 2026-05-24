#!/usr/bin/env python3
"""
Reads leads.json and writes/refreshes lead_scores.xlsx in the data/ folder.
Run after any lead is captured or updated.
"""

import json
import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip3 install openpyxl")

DATA_DIR = Path(__file__).parent / "data"
LEADS_FILE = DATA_DIR / "leads.json"
XLSX_FILE  = DATA_DIR / "lead_scores.xlsx"

SCORE_TIER_COLOURS = {
    "Very Hot": "FF4444",
    "Hot":      "FF9900",
    "Warm":     "FFDD00",
    "Cold":     "AAAAAA",
}

HEADER_FILL   = PatternFill("solid", fgColor="1F2D3D")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COLUMNS = [
    ("Lead ID",              18),
    ("Captured At",          20),
    ("Customer Name",        20),
    ("Phone",                16),
    ("Email",                28),
    ("Products of Interest", 35),
    ("Price (PKR)",          14),
    ("Intent",               18),
    ("Lead Score",           12),
    ("Score Tier",           12),
    ("Status",               12),
    ("Notes",                35),
]


def load_leads():
    if not LEADS_FILE.exists():
        return []
    with open(LEADS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("leads", [])


def format_products(products):
    if not products:
        return ""
    parts = []
    for p in products:
        if isinstance(p, dict):
            parts.append(p.get("name", str(p)))
        else:
            parts.append(str(p))
    return ", ".join(parts)


def total_price(products):
    if not products:
        return 0
    total = 0
    for p in products:
        if isinstance(p, dict):
            total += p.get("price_pkr", 0)
    return total


def build_sheet(ws, leads):
    # Headers
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, lead in enumerate(leads, start=2):
        products     = lead.get("products_of_interest", [])
        product_str  = format_products(products)
        price_total  = total_price(products)
        score        = lead.get("lead_score", 0)
        tier         = lead.get("score_tier", "Cold")
        captured_raw = lead.get("captured_at", "")

        try:
            captured_dt = datetime.fromisoformat(captured_raw.replace("Z", "+00:00"))
            captured_str = captured_dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            captured_str = captured_raw

        row_data = [
            lead.get("id", ""),
            captured_str,
            lead.get("name") or "—",
            lead.get("phone", ""),
            lead.get("email") or "—",
            product_str or "—",
            price_total if price_total else "—",
            lead.get("intent", ""),
            score,
            tier,
            lead.get("status", "new"),
            lead.get("notes") or lead.get("pain_point") or "—",
        ]

        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            # Alternate row shading (skip score-tier column which has its own colour)
            if col_idx != 10:
                cell.fill = fill

        # Colour the Score Tier cell
        tier_cell = ws.cell(row=row_idx, column=10)
        tier_colour = SCORE_TIER_COLOURS.get(tier, "FFFFFF")
        tier_cell.fill = PatternFill("solid", fgColor=tier_colour)
        tier_cell.font = Font(bold=True, color="FFFFFF" if tier in ("Very Hot", "Hot") else "333333")
        tier_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Bold the score number
        ws.cell(row=row_idx, column=9).font = Font(bold=True)
        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 22


def main():
    leads = load_leads()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Scores"

    build_sheet(ws, leads)

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    counts = {"Very Hot": 0, "Hot": 0, "Warm": 0, "Cold": 0}
    for lead in leads:
        tier = lead.get("score_tier", "Cold")
        counts[tier] = counts.get(tier, 0) + 1

    ws2.append(["Tier", "Count"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for tier, count in counts.items():
        ws2.append([tier, count])

    ws2.append([])
    ws2.append(["Total Leads", len(leads)])
    ws2.append(["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    for col in ("A", "B"):
        ws2.column_dimensions[col].width = 18

    wb.save(XLSX_FILE)
    print(f"[lead_scores] Saved {len(leads)} leads → {XLSX_FILE}")


if __name__ == "__main__":
    main()
